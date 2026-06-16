"""
openpyxl Excel generator — 7-sheet institutional workbook.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers,
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from app.core.reports.data_collector import ReportMetrics

# ── Colour constants (ARGB) ────────────────────────────────────────────────────

NAVY    = "FF1E3A5F"
ACCENT  = "FF00D4FF"
GOLD    = "FFCC9900"
WHITE   = "FFFFFFFF"
OFFWH   = "FFF0F4FF"
LIGHT   = "FFF8F9FA"
GREY    = "FF6C757D"
DARKGR  = "FF343A40"
GREEN   = "FF28A745"
RED     = "FFDC3545"
AMBER   = "FFFFC107"
BORDER  = "FFDEE2E6"


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def _font(bold=False, size=10, color=DARKGR, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _border(style="thin", color=BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_row(ws, row: int, headers: List[str], widths: Optional[List[int]] = None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE, size=9)
        cell.alignment = _align("center")
        cell.border = _border("thin", NAVY)
        if widths and col - 1 < len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def _value_cell(ws, row: int, col: int, value, fmt=None, bold=False, align="left",
                bg=None, color=DARKGR):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=9, color=color)
    cell.alignment = _align(align)
    cell.border = _border("hair", BORDER)
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = _fill(bg)
    return cell


def _section_title(ws, row: int, title: str, cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _fill(ACCENT)
    cell.font = _font(bold=True, size=10, color=NAVY)
    cell.alignment = _align("left")


def _add_cover_info(ws, m: ReportMetrics):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws["A1"] = "QUANTDESK — INSTITUTIONAL REPORT"
    ws["A1"].font = _font(bold=True, size=14, color=NAVY)
    ws["A2"] = m.report_name
    ws["A2"].font = _font(bold=True, size=12, color=DARKGR)
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"].font = _font(size=9, color=GREY, italic=True)
    ws["A4"] = f"Tickers: {', '.join(m.tickers)}"
    ws["A4"].font = _font(size=9, color=DARKGR)
    ws["A5"] = f"Period: {m.start_date} to {m.end_date}"
    ws["A5"].font = _font(size=9, color=DARKGR)
    ws["A6"] = "CONFIDENTIAL — For Authorized Recipients Only"
    ws["A6"].font = _font(size=8, color=GREY, italic=True)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet1_trade_history(wb: Workbook, m: ReportMetrics):
    ws = wb.active
    ws.title = "Trade History"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    _add_cover_info(ws, m)
    ws.row_dimensions[7].height = 8

    headers = ["Trade ID", "Symbol", "Entry Date", "Exit Date",
               "Entry Price", "Exit Price", "Position Size", "P&L ($)", "Return %", "Holding Period"]
    _header_row(ws, 8, headers, [12, 9, 13, 13, 12, 12, 14, 12, 11, 15])

    trades = m.trade_log or []
    if not trades:
        # Generate synthetic daily trade log from returns
        daily = m.daily_returns
        if not daily.empty:
            dates = daily.index
            for i, (date, ret) in enumerate(zip(dates[:-1:5], daily.iloc[:-1:5]), 1):
                exit_date = dates[min(i * 5, len(dates) - 1)]
                entry_p = 100.0
                exit_p = entry_p * (1 + ret * 5)
                pl = (exit_p - entry_p) * 1000
                trades.append({
                    "id": f"T{i:05d}",
                    "symbol": m.tickers[i % len(m.tickers)] if m.tickers else "N/A",
                    "entry_date": date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date),
                    "exit_date": exit_date.strftime("%Y-%m-%d") if hasattr(exit_date, "strftime") else str(exit_date),
                    "entry_price": round(entry_p, 2),
                    "exit_price": round(exit_p, 2),
                    "size": 1000,
                    "pnl": round(pl, 2),
                    "return_pct": round(ret * 5 * 100, 2),
                    "holding": 5,
                })

    for r, trade in enumerate(trades[:500], 9):
        bg = LIGHT if r % 2 == 0 else WHITE
        vals = [
            trade.get("id", f"T{r:05d}"),
            trade.get("symbol", ""),
            trade.get("entry_date", ""),
            trade.get("exit_date", ""),
            trade.get("entry_price", 0),
            trade.get("exit_price", 0),
            trade.get("size", 0),
            trade.get("pnl", 0),
            trade.get("return_pct", 0) / 100 if abs(trade.get("return_pct", 0)) > 1 else trade.get("return_pct", 0),
            trade.get("holding", 0),
        ]
        fmts = [None, None, None, None, '"$"#,##0.00', '"$"#,##0.00',
                "#,##0", '"$"#,##0.00', '0.00%', '0 "days"']
        for col, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = _value_cell(ws, r, col, v, fmt=fmt, bg=bg)
            # Colour PnL column
            if col == 8 and isinstance(v, (int, float)):
                cell.font = _font(size=9, color=GREEN if v >= 0 else RED, bold=True)

    if len(trades) > 1:
        pnl_col = get_column_letter(8)
        last_row = 8 + len(trades)
        ws.conditional_formatting.add(
            f"{pnl_col}9:{pnl_col}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F85149",
                mid_type="num", mid_value=0, mid_color="FFFFFF",
                end_type="max", end_color="3FB950",
            ),
        )


def _sheet2_portfolio_stats(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Portfolio Statistics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18

    def pct(v): return v  # store as float, format as %
    def fp(v): return v

    _section_title(ws, 1, "PORTFOLIO OVERVIEW", 4)

    overview = [
        ("Report Name", m.report_name, "Portfolio Value", f"${m.portfolio_value:,.0f}"),
        ("Tickers", ", ".join(m.tickers[:8]), "Benchmark", m.benchmark),
        ("Start Date", m.start_date, "End Date", m.end_date),
        ("Risk Score", f"{m.risk_score}/100", "Risk Label", m.risk_label),
    ]
    for r, (l1, v1, l2, v2) in enumerate(overview, 2):
        ws.cell(r, 1, l1).font = _font(bold=True, size=9, color=GREY)
        ws.cell(r, 2, v1).font = _font(size=9)
        ws.cell(r, 3, l2).font = _font(bold=True, size=9, color=GREY)
        ws.cell(r, 4, v2).font = _font(size=9)

    ws.row_dimensions[6].height = 8
    _section_title(ws, 7, "RETURN METRICS", 4)

    metrics = [
        ("Total Return", m.total_return, "0.00%"),
        ("CAGR (Ann.)", m.cagr, "0.00%"),
        ("Best Day", m.best_day, "0.00%"),
        ("Worst Day", m.worst_day, "0.00%"),
        ("Avg Daily Return", m.avg_daily_return, "0.000%"),
        ("Win Rate", m.win_rate, "0.00%"),
        ("Profit Factor", m.profit_factor, "0.00"),
    ]
    for r, (label, val, fmt) in enumerate(metrics, 8):
        bg = OFFWH if r % 2 == 0 else WHITE
        ws.cell(r, 1, label).font = _font(size=9, color=GREY)
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        c.font = _font(bold=True, size=9,
                       color=GREEN if isinstance(val, float) and val > 0 else
                       RED if isinstance(val, float) and val < -0.001 else DARKGR)
        c.fill = _fill(bg)

    ws.row_dimensions[15].height = 8

    # Individual asset performance table
    _section_title(ws, 16, "INDIVIDUAL ASSET PERFORMANCE", 4)
    _header_row(ws, 17, ["Ticker", "Weight", "Total Return", "Sharpe"], [12, 12, 16, 12])
    for r, t in enumerate(m.tickers[:30], 18):
        bg = OFFWH if r % 2 == 0 else WHITE
        ret = m.individual_returns.get(t, 0)
        _value_cell(ws, r, 1, t, bold=True, bg=bg)
        _value_cell(ws, r, 2, m.weights.get(t, 0), fmt="0.0%", bg=bg)
        c = _value_cell(ws, r, 3, ret, fmt="0.00%", bg=bg,
                         color=GREEN if ret > 0 else RED)
        _value_cell(ws, r, 4, m.individual_sharpe.get(t, 0), fmt="0.00", bg=bg)


def _sheet3_risk_metrics(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Risk Metrics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18

    _section_title(ws, 1, "RISK METRICS", 4)

    risk = [
        ("Annualized Volatility", m.volatility, "0.00%"),
        ("Downside Volatility", m.downside_vol, "0.00%"),
        ("Max Drawdown", m.max_drawdown, "0.00%"),
        ("DD Duration (days)", m.max_drawdown_duration, "0"),
        ("VaR (95%)", m.var_95, "0.000%"),
        ("CVaR / Expected Shortfall", m.cvar_95, "0.000%"),
        ("Risk Score", m.risk_score, "0"),
    ]
    ra = [
        ("Sharpe Ratio", m.sharpe, "0.00"),
        ("Sortino Ratio", m.sortino, "0.00"),
        ("Calmar Ratio", m.calmar, "0.00"),
        ("Information Ratio", m.information_ratio, "0.00"),
        ("Treynor Ratio", m.treynor, "0.00"),
        ("Alpha (Ann.)", m.alpha, "0.00%"),
        ("Beta", m.beta, "0.00"),
    ]

    _section_title(ws, 2, "Volatility & Tail Risk", 2)
    for r, (label, val, fmt) in enumerate(risk, 3):
        bg = OFFWH if r % 2 == 0 else WHITE
        ws.cell(r, 1, label).font = _font(size=9, color=GREY)
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        c.font = _font(bold=True, size=9)
        c.fill = _fill(bg)

    _section_title(ws, 2, "Risk-Adjusted Returns", 2)
    ws.cell(2, 3).fill = _fill(NAVY)
    ws.cell(2, 3).value = "Risk-Adjusted Returns"
    ws.cell(2, 3).font = _font(bold=True, size=10, color=WHITE)
    for r, (label, val, fmt) in enumerate(ra, 3):
        bg = OFFWH if r % 2 == 0 else WHITE
        ws.cell(r, 3, label).font = _font(size=9, color=GREY)
        c = ws.cell(r, 4, val)
        c.number_format = fmt
        c.font = _font(bold=True, size=9,
                       color=GREEN if isinstance(val, float) and val > 1.0 else
                       RED if isinstance(val, float) and val < 0 else DARKGR)
        c.fill = _fill(bg)


def _sheet4_performance_metrics(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Performance Metrics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14

    _section_title(ws, 1, "ANNUAL RETURNS", 4)
    _header_row(ws, 2, ["Year", "Portfolio", "Benchmark", "Active Return"], [10, 14, 14, 15])

    annual = m.annual_returns
    bench = m.benchmark_annual

    for r, (date, val) in enumerate(annual.items(), 3):
        year = date.year if hasattr(date, "year") else str(date)[:4]
        bg = OFFWH if r % 2 == 0 else WHITE
        bval = bench.get(date, None) if bench is not None and not bench.empty else None
        active = val - bval if bval is not None else None

        _value_cell(ws, r, 1, year, bg=bg, align="center")
        c2 = _value_cell(ws, r, 2, val, fmt="0.00%", bg=bg, align="center",
                          color=GREEN if val > 0 else RED)
        if bval is not None:
            c3 = _value_cell(ws, r, 3, bval, fmt="0.00%", bg=bg, align="center",
                              color=GREEN if bval > 0 else RED)
        if active is not None:
            c4 = _value_cell(ws, r, 4, active, fmt="+0.00%;-0.00%;0.00%", bg=bg, align="center",
                              color=GREEN if active > 0 else RED)

    if not annual.empty:
        last = 2 + len(annual) + 1
        ws.cell(last, 1, "Average").font = _font(bold=True, size=9)
        avg_cell = ws.cell(last, 2)
        avg_cell.value = annual.mean()
        avg_cell.number_format = "0.00%"
        avg_cell.font = _font(bold=True, size=9, color=NAVY)
        avg_cell.fill = _fill(ACCENT)


def _sheet5_monthly_heatmap(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Monthly Returns")
    ws.sheet_view.showGridLines = False

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Annual"]

    _section_title(ws, 1, "MONTHLY RETURNS HEATMAP", 14)
    _header_row(ws, 2, ["Year"] + months,
                [8] + [8] * 13)

    tbl = m.monthly_returns_table
    if tbl.empty:
        return

    annual_by_year = m.daily_returns.resample("YE").apply(lambda x: (1 + x).prod() - 1)

    for r_i, year in enumerate(sorted(tbl.index), 3):
        bg = OFFWH if r_i % 2 == 0 else WHITE
        ws.cell(r_i, 1, year).font = _font(bold=True, size=9)
        ws.cell(r_i, 1).fill = _fill(LIGHT)
        ws.cell(r_i, 1).alignment = _align("center")

        for col_i in range(1, 13):
            val = tbl.loc[year, col_i] if col_i in tbl.columns else None
            c = ws.cell(r_i, col_i + 1)
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                c.value = val
                c.number_format = "0.0%"
            else:
                c.value = None
            c.alignment = _align("center")
            c.font = _font(size=8.5)
            c.border = _border("hair", BORDER)

        # Annual column
        try:
            ann_idx = [i for i, d in enumerate(annual_by_year.index)
                       if (d.year if hasattr(d, "year") else int(str(d)[:4])) == year]
            ann_val = annual_by_year.iloc[ann_idx[0]] if ann_idx else None
        except Exception:
            ann_val = None

        ann_cell = ws.cell(r_i, 14)
        if ann_val is not None:
            ann_cell.value = ann_val
            ann_cell.number_format = "0.0%"
            ann_cell.font = _font(bold=True, size=9,
                                   color=GREEN if ann_val > 0 else RED)
        ann_cell.fill = _fill(LIGHT)
        ann_cell.alignment = _align("center")

    # Color scale for monthly data
    last_row = 2 + len(tbl)
    ws.conditional_formatting.add(
        f"B3:{get_column_letter(13)}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F85149",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="3FB950",
        ),
    )


def _sheet6_factor_exposure(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Factor Exposure")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14

    _section_title(ws, 1, "FACTOR EXPOSURE ANALYSIS", 4)

    if m.factor_exposures:
        _header_row(ws, 2, ["Factor", "Exposure"], [26, 14])
        for r, (factor, val) in enumerate(m.factor_exposures.items(), 3):
            bg = OFFWH if r % 2 == 0 else WHITE
            _value_cell(ws, r, 1, factor, bg=bg)
            c = _value_cell(ws, r, 2, val, fmt="0.00", bg=bg, align="center",
                             color=GREEN if val > 0 else RED if val < 0 else GREY)
    else:
        ws.cell(3, 1, "No factor data available. Run the Strategy Builder to generate factor exposures.").font = \
            _font(size=9, color=GREY, italic=True)

    # Correlation matrix
    if not m.correlation_matrix.empty:
        n = len(m.correlation_matrix)
        start_row = max(6, 4 + (len(m.factor_exposures) if m.factor_exposures else 2))
        _section_title(ws, start_row, "CORRELATION MATRIX", n + 1)
        _header_row(ws, start_row + 1,
                    [""] + list(m.correlation_matrix.columns),
                    [14] + [10] * n)
        for r, (idx, row) in enumerate(m.correlation_matrix.iterrows(), start_row + 2):
            ws.cell(r, 1, str(idx)).font = _font(bold=True, size=9)
            ws.cell(r, 1).fill = _fill(LIGHT)
            for c, val in enumerate(row, 2):
                cell = ws.cell(r, c, round(float(val), 3))
                cell.number_format = "0.00"
                cell.alignment = _align("center")
                cell.font = _font(size=8.5, bold=(abs(val) > 0.7 and c - 2 != r - start_row - 2))

        corr_range = f"B{start_row+2}:{get_column_letter(n+1)}{start_row+1+n}"
        ws.conditional_formatting.add(
            corr_range,
            ColorScaleRule(
                start_type="num", start_value=-1, start_color="F85149",
                mid_type="num", mid_value=0, mid_color="FFFFFF",
                end_type="num", end_value=1, end_color="3FB950",
            ),
        )


def _sheet7_attribution(wb: Workbook, m: ReportMetrics):
    ws = wb.create_sheet("Strategy Attribution")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    _section_title(ws, 1, "STRATEGY ATTRIBUTION", 5)

    row = 2
    _section_title(ws, row, "RISK DECOMPOSITION", 5)
    row += 1
    _header_row(ws, row, ["Metric", "Value", "vs Benchmark", "Percentile", "Assessment"],
                [28, 14, 18, 14, 22])
    row += 1

    def assess(ratio, thresholds, labels):
        for thresh, label in zip(thresholds, labels):
            if ratio >= thresh:
                return label
        return labels[-1]

    attrib = [
        ("Sharpe Ratio", m.sharpe, 1.0, "0.00",
         assess(m.sharpe, [1.5, 1.0, 0.5], ["Excellent", "Good", "Average", "Poor"])),
        ("Sortino Ratio", m.sortino, 1.0, "0.00",
         assess(m.sortino, [2.0, 1.2, 0.7], ["Excellent", "Good", "Average", "Poor"])),
        ("Calmar Ratio", m.calmar, 0.5, "0.00",
         assess(m.calmar, [1.5, 0.75, 0.4], ["Excellent", "Good", "Average", "Poor"])),
        ("Information Ratio", m.information_ratio, 0.5, "0.00",
         assess(m.information_ratio, [1.0, 0.5, 0.2], ["Top Quartile", "2nd Quartile", "3rd Quartile", "Bottom Quartile"])),
        ("Win Rate", m.win_rate, 0.5, "0.0%",
         assess(m.win_rate, [0.60, 0.52, 0.47], ["Strong", "Above Average", "Average", "Below Average"])),
        ("Max Drawdown", m.max_drawdown, -0.10, "0.00%",
         "Contained" if abs(m.max_drawdown) < 0.10 else "Acceptable" if abs(m.max_drawdown) < 0.20 else "Elevated"),
    ]

    for r, (label, val, bench_val, fmt, assessment) in enumerate(attrib, row):
        bg = OFFWH if r % 2 == 0 else WHITE
        _value_cell(ws, r, 1, label, bg=bg)
        c2 = _value_cell(ws, r, 2, val, fmt=fmt, bg=bg, align="center",
                          color=GREEN if isinstance(val, float) and val > 0 else
                          RED if isinstance(val, float) and val < -0.001 else DARKGR)
        _value_cell(ws, r, 3, bench_val, fmt=fmt, bg=bg, align="center")
        active = val - bench_val if isinstance(val, float) and isinstance(bench_val, float) else None
        if active is not None:
            _value_cell(ws, r, 4, active, fmt=fmt, bg=bg, align="center",
                         color=GREEN if active >= 0 else RED)
        ws.cell(r, 5, assessment).font = _font(size=9,
                                                color=GREEN if assessment in ("Excellent", "Strong", "Top Quartile", "Good", "Contained", "Above Average")
                                                else RED if assessment in ("Poor", "Bottom Quartile", "Elevated", "Below Average")
                                                else AMBER)
        ws.cell(r, 5).fill = _fill(bg)
        ws.cell(r, 5).alignment = _align("center")


# ── Main generator ─────────────────────────────────────────────────────────────

def generate_excel(m: ReportMetrics) -> bytes:
    wb = Workbook()
    _sheet1_trade_history(wb, m)
    _sheet2_portfolio_stats(wb, m)
    _sheet3_risk_metrics(wb, m)
    _sheet4_performance_metrics(wb, m)
    _sheet5_monthly_heatmap(wb, m)
    _sheet6_factor_exposure(wb, m)
    _sheet7_attribution(wb, m)

    # Global workbook properties
    wb.properties.title = m.report_name
    wb.properties.creator = "QuantDesk"
    wb.properties.description = "Institutional Research Report"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
